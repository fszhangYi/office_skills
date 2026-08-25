#!/usr/bin/env python3
"""Convert Markdown 立项预研稿 to 立项报告-style 2-column XLSX."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Layout aligned with 前盖具身智能上下料技术预研.xlsx
COL_A_WIDTH = 16.725
COL_B_WIDTH = 141.35
FONT_NAME = "微软雅黑"
FONT_SIZE = 11
MAX_ROW_HEIGHT = 400.0
LINE_PT = 16.0
PAD_PT = 12.0

THIN = Side(style="thin", color="FF000000")


def parse_md_sections(text: str) -> tuple[str | None, list[tuple[str, str]]]:
    """Return (doc_title, [(heading, body), ...]) for ## / ### headings."""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")
    doc_title = None
    sections: list[tuple[str, str]] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        m1 = re.match(r"^#\s+(.+)$", line)
        m2 = re.match(r"^(#{2,3})\s+(.+)$", line)
        if m1 and not line.startswith("##"):
            doc_title = m1.group(1).strip()
            i += 1
            continue
        if m2:
            level, title = m2.group(1), m2.group(2).strip()
            i += 1
            body_lines: list[str] = []
            while i < len(lines):
                if re.match(r"^#{1,3}\s+", lines[i]):
                    break
                body_lines.append(lines[i])
                i += 1
            body = "\n".join(body_lines).strip()
            # ## and ### both become rows (### keeps full title for A column)
            sections.append((title, body))
            continue
        i += 1
    return doc_title, sections


def md_table_to_list(block: str) -> str | None:
    """Convert a GFM table block to numbered prose list; else None."""
    lines = [ln.rstrip() for ln in block.strip().split("\n") if ln.strip()]
    if len(lines) < 2:
        return None
    if not all("|" in ln for ln in lines):
        return None
    # skip separator row
    rows = []
    for ln in lines:
        if re.match(r"^\|?\s*:?-+:?\s*(\|\s*:?-+:?\s*)+\|?$", ln):
            continue
        cells = [c.strip() for c in ln.strip().strip("|").split("|")]
        rows.append(cells)
    if len(rows) < 2:
        return None
    header = rows[0]
    out = []
    for idx, row in enumerate(rows[1:], 1):
        while len(row) < len(header):
            row.append("")
        if len(header) >= 4:
            # e.g. 策略 | 角色 | 保留/晋升 | 退出/降级
            name, role, keep, exit_ = row[0], row[1], row[2], row[3]
            piece = f"{idx}. {name}"
            if role:
                piece += f"（{role}）"
            if keep:
                piece += f"\n   保留/晋升：{keep}"
            if exit_:
                piece += f"\n   退出/降级：{exit_}"
            # extra columns if any
            for h, c in zip(header[4:], row[4:]):
                if c:
                    piece += f"\n   {h}：{c}"
            out.append(piece)
        elif len(header) == 3:
            dim, val, note = row[0], row[1], row[2]
            piece = f"{idx}. {dim}：{val}"
            if note:
                piece += f"；{note}"
            out.append(piece)
        elif len(header) == 2:
            out.append(f"{idx}. {row[0]}：{row[1]}")
        else:
            out.append(f"{idx}. " + "；".join(c for c in row if c))
    return "\n".join(out)


def transform_body_tables(body: str) -> str:
    """Replace GFM tables in body with numbered lists (Excel-readable)."""
    if "|" not in body:
        return body
    lines = body.split("\n")
    out: list[str] = []
    i = 0
    while i < len(lines):
        if "|" in lines[i]:
            start = i
            block = []
            while i < len(lines) and ("|" in lines[i] or not lines[i].strip()):
                if lines[i].strip():
                    block.append(lines[i])
                elif block:
                    # blank ends table
                    break
                i += 1
            converted = md_table_to_list("\n".join(block))
            if converted:
                out.append(converted)
            else:
                out.extend(block)
            continue
        out.append(lines[i])
        i += 1
    return "\n".join(out).strip()


def parse_date(value: str):
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return value


def count_wrapped_lines(text: str, chars_per_line: int) -> int:
    if not text:
        return 1
    lines = 0
    for para in str(text).split("\n"):
        if para == "":
            lines += 1
        else:
            lines += max(1, (len(para) + chars_per_line - 1) // chars_per_line)
    return max(1, lines)


def row_height(text, row_idx: int, chars_per_line: int) -> float:
    if row_idx <= 3 and (isinstance(text, (date, datetime)) or (isinstance(text, str) and len(text) < 40)):
        return {1: 15.0, 2: 16.5, 3: 16.5}.get(row_idx, 16.5)
    if isinstance(text, (date, datetime)):
        return 16.5
    n = count_wrapped_lines(str(text), chars_per_line)
    h = n * LINE_PT + PAD_PT
    if n <= 2 and len(str(text)) < 80:
        return max(28.0, min(h, 40.0))
    return min(MAX_ROW_HEIGHT, max(28.0, h))


def split_next_plan(body: str) -> tuple[str, str] | None:
    """If body starts with a one-line summary then details, split for merge layout."""
    parts = [p.strip() for p in re.split(r"\n\s*\n", body.strip()) if p.strip()]
    if len(parts) >= 2 and "\n" not in parts[0] and len(parts[0]) <= 80:
        return parts[0], "\n\n".join(parts[1:])
    # first line + rest
    lines = body.strip().split("\n")
    if len(lines) >= 2 and not re.match(r"^\s*\d+[\.、]", lines[0]):
        return lines[0].strip(), "\n".join(lines[1:]).strip()
    return None


def split_oversized_body(label: str, body: str, chars_per_line: int) -> list[tuple[str, str]]:
    """Split body into multiple (label, chunk) so each fits under MAX_ROW_HEIGHT."""
    if not body:
        return [(label, body)]
    needed = count_wrapped_lines(body, chars_per_line) * LINE_PT + PAD_PT
    if needed <= MAX_ROW_HEIGHT:
        return [(label, body)]

    # Prefer split on blank lines; fallback to hard line chunks
    paras = re.split(r"\n\s*\n", body.strip())
    chunks: list[str] = []
    buf: list[str] = []

    def buf_text() -> str:
        return "\n\n".join(buf).strip()

    for para in paras:
        trial = buf + [para]
        trial_text = "\n\n".join(trial).strip()
        h = count_wrapped_lines(trial_text, chars_per_line) * LINE_PT + PAD_PT
        if buf and h > MAX_ROW_HEIGHT:
            chunks.append(buf_text())
            buf = [para]
        else:
            buf.append(para)
    if buf:
        chunks.append(buf_text())

    # If still one huge para, split by lines
    fixed: list[str] = []
    for ch in chunks:
        h = count_wrapped_lines(ch, chars_per_line) * LINE_PT + PAD_PT
        if h <= MAX_ROW_HEIGHT:
            fixed.append(ch)
            continue
        lines = ch.split("\n")
        buf_l: list[str] = []
        for ln in lines:
            trial = "\n".join(buf_l + [ln])
            ht = count_wrapped_lines(trial, chars_per_line) * LINE_PT + PAD_PT
            if buf_l and ht > MAX_ROW_HEIGHT:
                fixed.append("\n".join(buf_l))
                buf_l = [ln]
            else:
                buf_l.append(ln)
        if buf_l:
            fixed.append("\n".join(buf_l))

    out: list[tuple[str, str]] = []
    for i, ch in enumerate(fixed):
        out.append((label if i == 0 else f"{label}（续{i}）", ch))
    return out


def md_to_xlsx(md_path: Path, xlsx_path: Path) -> None:
    text = md_path.read_text(encoding="utf-8")
    doc_title, sections = parse_md_sections(text)
    if not sections:
        raise SystemExit(f"error: no ## / ### sections in {md_path}")

    # Ensure 预研名称 row if missing but doc title exists
    labels_norm = [re.sub(r"\s+", "", s[0]) for s in sections]
    if doc_title and not any(x in {"预研名称", "项目名称", "立项名称"} for x in labels_norm):
        sections.insert(0, ("预研名称", f"《{doc_title}》"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = COL_B_WIDTH
    chars_per_line = max(36, int(COL_B_WIDTH / 2.6))

    # Expand oversized sections before writing
    expanded: list[tuple[str, str]] = []
    for label, body in sections:
        body = transform_body_tables(body)
        expanded.extend(split_oversized_body(label, body, chars_per_line))
    sections = expanded

    label_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FF000000")
    value_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=False, color="FF000000")
    value_font_bold = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FF000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def label_border():
        return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def value_border():
        return Border(left=None, right=THIN, top=THIN, bottom=THIN)

    bold_labels = {
        "技术升级",
        "当前能力",
        "研究目标",
        "研究内容",
        "主备退出标准",
        "导入硬门槛与ROI",
        "导入硬门槛与 ROI",
        "资源需求",
        "预研名称",
    }

    r = 1
    i = 0
    while i < len(sections):
        label, body = sections[i]
        norm = re.sub(r"\s+", "", label)
        # strip （续N） for special-case matching
        norm_base = re.sub(r"（续\d+）$", "", norm)

        # 建议与下一步计划 → merge A across 2 rows when splittable
        if "建议与下一步" in norm_base or norm_base in {"下一步计划", "建议"}:
            split = split_next_plan(body) if body else None
            a1, b1 = ws.cell(r, 1), ws.cell(r, 2)
            a1.value = label
            a1.font = label_font
            a1.alignment = center
            a1.border = label_border()
            if split:
                title, detail = split
                b1.value = title
                b1.font = value_font
                b1.alignment = left
                b1.border = value_border()
                ws.row_dimensions[r].height = row_height(title, r, chars_per_line)
                r2 = r + 1
                b2 = ws.cell(r2, 2)
                b2.value = detail
                b2.font = value_font
                b2.alignment = left
                b2.border = value_border()
                ws.cell(r2, 1).border = label_border()
                ws.cell(r2, 1).font = label_font
                ws.cell(r2, 1).alignment = center
                ws.merge_cells(start_row=r, start_column=1, end_row=r2, end_column=1)
                ws.row_dimensions[r2].height = row_height(detail, r2, chars_per_line)
                r = r2 + 1
            else:
                b1.value = body
                b1.font = value_font
                b1.alignment = left
                b1.border = value_border()
                ws.row_dimensions[r].height = row_height(body, r, chars_per_line)
                r += 1
            i += 1
            continue

        a, b = ws.cell(r, 1), ws.cell(r, 2)
        # Multi-line A for long compound labels (ignore 续N suffix for layout)
        if norm_base == "安全与容错机制" and "续" not in label:
            a.value = " 安全与\n容错机制"
        elif norm_base in {"放置允差与纸板处理", "放置允差与节拍基线"} and "续" not in label:
            a.value = label.replace("与", "与\n", 1) if "与" in label and "\n" not in label else label
        elif norm_base == "节拍基线与灵活启停" and "续" not in label:
            a.value = "节拍基线与\n灵活启停"
        elif norm_base in {"导入硬门槛与ROI", "导入硬门槛与ROI"} or (
            norm_base.replace(" ", "") == "导入硬门槛与ROI" and "续" not in label
        ):
            a.value = "导入硬门槛\n与 ROI" if "续" not in label else label
        else:
            a.value = label

        # date field
        if norm_base == "提出日期" and body:
            b.value = parse_date(body.split("\n")[0].strip())
            if isinstance(b.value, date):
                b.number_format = "YYYY-MM-DD"
        elif norm_base in {"预研名称", "项目名称", "立项名称"} and body and not body.startswith("《"):
            b.value = f"《{body}》" if doc_title and body == doc_title else body
        else:
            b.value = body

        a.font = label_font
        a.alignment = center
        a.border = label_border()
        use_bold = norm_base in bold_labels or any(
            norm_base.startswith(bl) for bl in ("研究内容", "研究目标", "技术升级")
        )
        b.font = value_font_bold if use_bold else value_font
        if r == 1 and norm_base in {"预研名称", "项目名称", "立项名称"}:
            b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            b.alignment = left
        b.border = value_border()
        ws.row_dimensions[r].height = row_height(b.value, r, chars_per_line)
        r += 1
        i += 1

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="立项预研 Markdown → 立项报告 XLSX")
    p.add_argument("input", type=Path, help="input .md")
    p.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="output .xlsx (default: <stem> 立项报告.xlsx or same stem .xlsx)",
    )
    args = p.parse_args(argv)
    inp = args.input
    if not inp.is_file():
        print(f"error: not found: {inp}", file=sys.stderr)
        return 1
    out = args.output
    if out is None:
        out = inp.with_name(f"{inp.stem}.xlsx")
    try:
        md_to_xlsx(inp, out)
    except SystemExit as e:
        print(e, file=sys.stderr)
        return 1
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
