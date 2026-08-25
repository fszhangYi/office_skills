#!/usr/bin/env python3
"""Convert 立项报告-style 2-column XLSX to Markdown."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def _norm_label(label: str) -> str:
    return re.sub(r"\s+", "", label.replace("\n", "")).strip()


def _is_title_like(label: str) -> bool:
    return _norm_label(label) in {"预研名称", "项目名称", "立项名称"}


def iter_label_value_rows(ws):
    """Yield (label, value) in row order; expand merged A labels."""
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col != 1 or mr.max_col != 1:
            continue
        top = ws.cell(mr.min_row, 1).value
        for r in range(mr.min_row, mr.max_row + 1):
            merged_map[r] = top

    max_row = ws.max_row or 0
    r = 1
    while r <= max_row:
        a_cell = ws.cell(r, 1)
        b_cell = ws.cell(r, 2)
        if isinstance(a_cell, MergedCell):
            label = merged_map.get(r)
        else:
            label = a_cell.value if a_cell.value is not None else merged_map.get(r)
        label_s = _cell_str(label)
        value_s = _cell_str(b_cell.value)

        # Consume consecutive rows that share the same merged label
        if r in merged_map:
            parts = []
            label_s = _cell_str(merged_map[r])
            start = r
            while r <= max_row and merged_map.get(r) == merged_map[start]:
                v = _cell_str(ws.cell(r, 2).value)
                if v:
                    parts.append(v)
                r += 1
            yield label_s, "\n\n".join(parts)
            continue

        if not label_s and not value_s:
            r += 1
            continue
        yield label_s, value_s
        r += 1


def xlsx_to_md(xlsx_path: Path) -> str:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    blocks: list[str] = []
    doc_title = None
    sections: list[tuple[str, str]] = []

    for label, value in iter_label_value_rows(ws):
        if not label and not value:
            continue
        if _is_title_like(label) and value:
            doc_title = value.strip("《》")
            sections.append((label, value))
            continue
        sections.append((label or "未命名", value))

    if doc_title:
        blocks.append(f"# {doc_title}")
        blocks.append("")
    elif sections:
        # fallback: use first section value or workbook stem
        blocks.append(f"# {xlsx_path.stem}")
        blocks.append("")

    for label, value in sections:
        heading = _norm_label(label) if label else "未命名"
        # keep readable breaks for multi-line labels already normalized
        if label and "\n" in str(label):
            heading = re.sub(r"\s+", "", str(label).replace("\n", ""))
        blocks.append(f"## {heading}")
        blocks.append("")
        if value:
            blocks.append(value)
            blocks.append("")

    return "\n".join(blocks).rstrip() + "\n"


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="立项报告 XLSX → Markdown")
    p.add_argument("input", type=Path, help="input .xlsx")
    p.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="output .md (default: same stem next to input)",
    )
    args = p.parse_args(argv)
    inp = args.input
    if not inp.is_file():
        print(f"error: not found: {inp}", file=sys.stderr)
        return 1
    out = args.output or inp.with_suffix(".md")
    text = xlsx_to_md(inp)
    out.write_text(text, encoding="utf-8")
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
